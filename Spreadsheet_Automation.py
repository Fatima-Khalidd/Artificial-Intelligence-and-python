import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row, 13)
        corrected_marks = cell.value + 50
        corrected_marks_cell = sheet.cell(row, 14)
        corrected_marks_cell.value = corrected_marks

    values = Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=15,
                        max_col=15)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "n2")

    wb.save(filename)